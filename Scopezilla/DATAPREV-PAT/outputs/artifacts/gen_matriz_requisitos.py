#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera a Matriz Comparativa de Requisitos — Documento de Visão DM.210824 v0.7 (MTE/Dataprev)
vs. escopo Salesforce PS (epics.json E01-E10, ADRs 0001-0006).

Requisitos transcritos FIELMENTE do PDF. Coluna comparativa:
  - Integral : consideramos por inteiro no escopo atual
  - Parcial  : consideramos com ressalva / divergência de premissa
  - Não      : fora do escopo atual / frente nova / conflito
Não altera data/*.json — artefato de análise autônomo.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- paleta ----
NAVY   = "032D60"   # Salesforce navy
BLUE   = "0176D3"
GREEN  = "C7E5C0"   # integral
YELLOW = "FCE8B2"   # parcial
RED    = "F5C9C4"   # não
GREY   = "F3F3F3"
WHITE  = "FFFFFF"

hdr_font   = Font(name="Calibri", bold=True, color=WHITE, size=11)
grp_font   = Font(name="Calibri", bold=True, color=WHITE, size=11)
cell_font  = Font(name="Calibri", size=10)
id_font    = Font(name="Calibri", bold=True, size=10)
hdr_fill   = PatternFill("solid", fgColor=NAVY)
grp_fill   = PatternFill("solid", fgColor=BLUE)
thin       = Side(style="thin", color="D0D0D0")
border     = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap_top   = Alignment(wrap_text=True, vertical="top")
wrap_ctr   = Alignment(wrap_text=True, vertical="center", horizontal="center")

STATUS_FILL = {"Integral": GREEN, "Parcial": YELLOW, "Não": RED}

HEADERS = ["ID", "Requisito (verbatim — Documento de Visão v0.7)", "Ator / Origem",
           "Consideramos?", "Como consideramos (épica + abordagem)", "Observação / Divergência"]
COLW    = [10, 62, 22, 15, 52, 56]


def style_header(ws, row=1):
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = wrap_ctr; cell.border = border
    for c, w in enumerate(COLW, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = f"A{row+1}"


def add_group(ws, r, text):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(HEADERS))
    cell = ws.cell(row=r, column=1, value=text)
    cell.font = grp_font; cell.fill = grp_fill
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
    cell.border = border
    ws.row_dimensions[r].height = 20
    return r + 1


def add_row(ws, r, rid, req, ator, status, como, obs):
    vals = [rid, req, ator, status, como, obs]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = border
        cell.alignment = wrap_ctr if c in (1, 4) else wrap_top
        cell.font = id_font if c == 1 else cell_font
        if c == 4:
            cell.fill = PatternFill("solid", fgColor=STATUS_FILL.get(status, WHITE))
            cell.font = Font(name="Calibri", bold=True, size=10)
    return r + 1


wb = Workbook()

# ===================================================================
# ABA 1 — REQUISITOS FUNCIONAIS (3.1 Necessidades e Funcionalidades)
# ===================================================================
ws = wb.active
ws.title = "Requisitos Funcionais"
style_header(ws)
r = 2

r = add_group(ws, r, "N01 — Marketplace: Cotação de Facilitadoras por Beneficiárias e Contratos Avulsos  ·  Permitir a intermediação entre empresas beneficiárias (PAT ou não) e facilitadoras, com processo de cotação padronizado, e registro de contratos firmados fora da plataforma.")
r = add_row(ws, r, "F01.01", "Criar pedido de cotação padronizado, solicitando às empresas, beneficiárias do PAT ou não PAT, os dados que serão disponibilizados para as Facilitadoras", "Beneficiária", "Integral", "E02 — demanda cadastrada como Opportunity nativa (ADR 0004): nº funcionários, valor, vigência, distribuição por UF, recursos obrigatórios.", "")
r = add_row(ws, r, "F01.02", "Disponibilizar cotação para Facilitadoras habilitadas", "Sistema", "Integral", "E02 + E05 — na Fase 1 a plataforma expõe um endpoint de consulta (pull via API/MuleSoft); a facilitadora consulta as demandas abertas na vigência.", "Notificação ativa (push) fica no roadmap futuro, com canal a definir (G0209/G0211). Na Fase 1 é pull, não push.")
r = add_row(ws, r, "F01.03", "Receber e armazenar proposta das Facilitadoras. Quando a Facilitadora ofertar mais de uma proposta, uma deve ser marcada como a principal", "Sistema/Facilitadora", "Integral", "E02 — N facilitadoras respondem com Quotes nativas via API (sem UI/licença). Marcação de proposta principal na camada custom de comparação.", "")
r = add_row(ws, r, "F01.04", "Possibilitar que a Beneficiária possa comparar as propostas recebidas e selecionar uma proposta vencedora. Ao selecionar uma proposta vencedora, o sistema deve apresentar os dados de contato com a Facilitadora", "Beneficiária", "Integral", "E02 — tela LWC 'Comparar Propostas'; seleção TRAVADA até o fechamento da janela de vigência (não seleção cega). Dados de contato liberados na seleção.", "")
r = add_row(ws, r, "F01.05", "Armazenar o contrato vencedor após assinatura, relacionando o com a proposta apresentada na plataforma", "Sistema/Facilitadora", "Integral", "E02 — upload do PDF imutável do contrato + metadados + versões. SEM CLM na Fase 1 (aditivos/renovações = upload de nova versão).", "")
r = add_row(ws, r, "F01.06", "Possibilitar a visualização do histórico de cotações, propostas e contratos", "Beneficiária", "Integral", "E02 — histórico sobre objetos nativos (Opportunity/Quote) + contratos armazenados.", "")
r = add_row(ws, r, "F01.07", "Possibilitar que as Facilitadoras possam enviar os contratos firmados fora da plataforma do Marketplace, inclusive os contratos firmados anteriores a plataforma", "Facilitadora", "Integral", "E02 — carga de contratos avulsos e legados (PAT e não-PAT), premissa 31/jul.", "")

r = add_group(ws, r, "N02 — Marketplace: Financeiro (Folha de Pagamento + Repasse)  ·  Permitir a gestão mensal da folha de benefícios e disponibilização do boleto.")
r = add_row(ws, r, "F02.01", "Envio da folha de benefícios e demais arquivos relacionados em formato padronizado a partir do contrato estabelecido", "Beneficiária", "Integral", "E03 — upload do CSV de folha da competência via portal ou API; validação de layout/integridade.", "")
r = add_row(ws, r, "F02.02", "Registrar a folha no sistema, com trabalhadores e valores", "Sistema", "Parcial", "E03 — registra o CABEÇALHO/competência da folha e valida o arquivo; habilita o download pela facilitadora.", "DIVERGÊNCIA: as LINHAS por trabalhador NÃO são persistidas em objeto na Fase 1 (só cabeçalho). O documento pede 'trabalhadores e valores' registrados — carga linha-a-linha é roadmap futuro. Ponto a alinhar.")
r = add_row(ws, r, "F02.03", "Disponibilizar consulta das folhas pendentes (polling) por Facilitadora e contrato", "Sistema/Facilitadora", "Integral", "E03 — arquivo habilitado para download da facilitadora por contrato/vigência; consulta via API (polling).", "")
r = add_row(ws, r, "F02.04", "Registrar os dados do processamento da folha realizado pela Facilitadora, possibilitando que a Beneficiária possa acompanhar o andamento e baixar o boleto para pagamento", "Sistema/Facilitadora/Beneficiária", "Integral", "E03 — facilitadora devolve status 'processado' + valor via API; plataforma aciona gateway, recebe boleto registrado e o disponibiliza à beneficiária no portal.", "")
r = add_row(ws, r, "F02.05", "Fazer upload da Nota Fiscal para a plataforma, com possibilidade de consulta pela Beneficiária", "Sistema/Facilitadora/Beneficiária", "Parcial", "E03 — fluxo financeiro (boleto/split) contemplado; a NF pode ser anexada ao fluxo documental do contrato.", "A CONFIRMAR: upload/consulta de Nota Fiscal não está explícito no escopo atual de E03 — extensão pequena, mas não dimensionada como item próprio.")

r = add_group(ws, r, "N03 — Marketplace: Integração com Sistema de Adesão do PAT (MTE)  ·  Integrar com o sistema existente, mantido pela DTI do MTE, que gerencia a adesão de Beneficiárias e Facilitadoras ao programa do PAT.")
r = add_row(ws, r, "F03.01", "Consultar adesão ativa de empresa Beneficiária", "Sistema (API)", "Integral", "E05 — integração MuleSoft com o Novo PAT (consulta de adesão da beneficiária).", "Novo PAT HOJE NÃO TEM API (MT/DTI) → mock-first obrigatório até os Swaggers serem disponibilizados.")
r = add_row(ws, r, "F03.02", "Consultar registro ativo de Facilitadora", "Sistema (API)", "Integral", "E05 — integração MuleSoft com o Novo PAT (consulta de registro da facilitadora).", "Idem F03.01 — dependência de API do Novo PAT (DTI/MTE).")

r = add_group(ws, r, "N04 — Credenciamento do Estabelecimento  ·  Permitir que restaurantes, supermercados e similares se credenciem junto ao governo para aceitar pagamentos via PAT, substituindo o processo hoje feito pelas facilitadoras.")
r = add_row(ws, r, "F04.01", "Solicitar credenciamento de estabelecimento comercial", "Estabelecimento", "Integral", "E04 — auto-credenciamento via gov.br PJ; face do estabelecimento que cadastra e sobe documentos.", "")
r = add_row(ws, r, "F04.02", "Validar dados do estabelecimento (situação cadastral do CNPJ junto a receita, atividade econômica (CNAE), regularidade sanitária)", "Sistema", "Integral", "E04 — validação automatizada com transbordo humano; vigilância sanitária: 5000+ padrões municipais, validade como parâmetro mínimo + extração por IA.", "")
r = add_row(ws, r, "F04.03", "Consultar status do credenciamento", "Estabelecimento", "Integral", "E04 — consulta de status pela face do estabelecimento.", "")
r = add_row(ws, r, "F04.04", "Disponibilizar lista de estabelecimentos credenciados ou descredenciados para facilitadoras/adquirentes", "Facilitadora/Adquirente", "Integral", "E04 + E05 — a adquirente consulta a base de estabelecimentos via API (client credentials, escopo restrito, rate limit/cache) antes de processar transações.", "")
r = add_row(ws, r, "F04.05", "Importar dados legados, referentes aos estabelecimentos já credenciados", "Sistema/NOVO PAT", "Integral", "E07 (Migração) + E04 — carga inicial de estabelecimentos credenciados a partir do Novo PAT/bases MTE.", "Volume de migração 700–800 mil estabelecimentos (Risco 5 do documento) — negociar absorção de base.")

r = add_group(ws, r, "N05 — Módulo de Gestão (MTE)  ·  Fornecer ao MTE um sistema que permita acompanhar e administrar todas as operações da plataforma PAT: credenciamento, cotações, contratações, financeiro e repasses.")
r = add_row(ws, r, "F05.01", "Consultar cotações em aberto (quem publicou, quando, volume)", "MTE", "Parcial", "E10 (Painel MTE) + views/relatórios nativos do Salesforce na org dedicada (E08).", "Não há uma épica 'Módulo de Gestão MTE' separada; a consulta é atendida por Tableau (VISÃO MTE) + UI nativa. Escopo do módulo administrativo a delimitar.")
r = add_row(ws, r, "F05.02", "Consultar propostas recebidas por cotação (facilitadoras, condições)", "MTE", "Parcial", "E10 + views/relatórios nativos sobre Quotes (E02).", "Idem F05.01 — cobertura por nativo + Tableau, sem módulo administrativo dedicado dimensionado.")
r = add_row(ws, r, "F05.03", "Consultar contratos (beneficiária ↔ facilitadora, vigência)", "MTE", "Parcial", "E10 + views/relatórios nativos sobre contratos (E02).", "Idem F05.01.")
r = add_row(ws, r, "F05.04", "Painel financeiro: folhas processadas, boletos e notas fiscais disponibilizadas", "MTE", "Parcial", "E10 (VISÃO NEGÓCIO/MTE — valores repassados, MDR médio) sobre E03; UI nativa complementar.", "Painel de acompanhamento coberto por Tableau; a granularidade financeira operacional (por boleto/NF) na UI de gestão a confirmar.")
r = add_row(ws, r, "F05.05", "Análise dos Estabelecimentos Credenciados. Ações de Credenciar e descredenciar Estabelecimentos", "MTE", "Integral", "E04 — face do Analista MTE: análise documento a documento (Válido/Inválido + motivo), parecer (Deferido/Exigência/Indeferido), trilha de auditoria, ações de credenciar/descredenciar.", "")

r = add_group(ws, r, "N06 — Monitoramento  ·  API para receber transações de vendas das Facilitadoras, permitindo ao MTE acompanhar as transações de venda realizadas utilizando auxílio alimentação ou refeição. Os dados de transações recebidas são a posteriori.")
r = add_row(ws, r, "F06.01", "API para recebimento de transações de vendas. Deve constar os dados da transação que possuem, como por exemplo, CPF do trabalhador, CNPJ do Estabelecimento, Valor, Data e Hora, Tipo auxílio (Alimentação/Refeição), IP/Porta, IMEI, Saldo, Taxa MDR, CNPJ Rede/Instituidora, CNPJ Adquirente, CNPJ da Beneficiária, CNPJ Facilitadora, entre outros", "Sistema/Facilitadora", "Não", "E05 pode expor o endpoint de ingestão (near-term); porém o MÓDULO DE MONITORAMENTO (processamento, armazenamento, analytics) NÃO está dimensionado na Fase 1.", "MUDANÇA-CHAVE: a proposta declarava Monitoramento/analytics 'fora desta rodada'. O documento coloca-o no MVP (15/nov), com ingestão de até 30M transações/dia — exige camada de staging/streaming (Data Cloud / fila MuleSoft / Heroku). Frente NOVA a dimensionar.")

r = add_group(ws, r, "N07 — CTPS Digital  ·  Notificações para que o trabalhador acompanhe o envio do crédito pela empresa beneficiária.")
r = add_row(ws, r, "F07.01", "Notificar trabalhador sobre crédito programado/enviado", "Trabalhador (via CTPS Digital)", "Integral", "E03 + E05 — 'expectativa de crédito' via integração com a CTPS Digital no processamento da folha (só monitoramento/notificação na Fase 1).", "")
r = add_row(ws, r, "F07.02", "Consulta pelo trabalhador do histórico de créditos recebidos", "Trabalhador (via CTPS Digital)", "Integral", "E05 — a plataforma alimenta a CTPS Digital; a consulta do histórico ocorre no app CTPS Digital (produto do governo).", "A UI de consulta é da CTPS Digital (cliente); nosso escopo é a integração/alimentação do dado.")

# ===================================================================
# ABA 2 — ARQUIVOS DE DADOS (3.2)
# ===================================================================
ws2 = wb.create_sheet("Arquivos de Dados")
style_header(ws2)
r = 2
r = add_group(ws2, r, "N01 — Contratação")
r = add_row(ws2, r, "AD01.01", "Cotações (pedidos de cotação das beneficiárias).", "—", "Integral", "E02 — Opportunity nativa.", "")
r = add_row(ws2, r, "AD01.02", "Propostas (respostas das facilitadoras).", "—", "Integral", "E02 — Quote nativa (via API).", "")
r = add_row(ws2, r, "AD01.03", "Contratos (registro de contratos ativos, PAT e não-PAT)", "—", "Integral", "E02 — contrato (PDF imutável + metadados + versões); PAT e não-PAT com regras distintas.", "")
r = add_group(ws2, r, "N02 — Financeiro")
r = add_row(ws2, r, "AD02.01", "Folhas de Pagamento (formato CSV padronizado).", "—", "Parcial", "E03 — cabeçalho/competência da folha; validação do CSV.", "Linhas por trabalhador não persistidas na Fase 1 (ver F02.02).")
r = add_row(ws2, r, "AD02.02", "Transações Financeiras (boletos, pagamentos, repasses, split).", "—", "Integral", "E03 — motor de regras de split, boletagem e conciliação; registra o racional (datas, split, ordens/boletagens).", "Fronteira ADR 0003: a EXECUÇÃO das transações bancárias e a custódia são do GATEWAY externo, fora do Salesforce.")
r = add_group(ws2, r, "N03 — NOVO PAT")
r = add_row(ws2, r, "AD03.01", "Consultar dados de Beneficiárias.", "—", "Integral", "E05 — consulta via API (MuleSoft) ao Novo PAT.", "Dependência de API do Novo PAT (DTI/MTE) — mock-first.")
r = add_row(ws2, r, "AD03.02", "Consultar dados de Facilitadoras.", "—", "Integral", "E05 — consulta via API (MuleSoft) ao Novo PAT.", "Idem AD03.01.")
r = add_group(ws2, r, "N04 — Credenciamento")
r = add_row(ws2, r, "AD04.01", "Estabelecimentos Credenciados.", "—", "Integral", "E04 — base de estabelecimentos credenciados; E07 carga inicial.", "")
r = add_group(ws2, r, "N06 — Monitoramento")
r = add_row(ws2, r, "AD06.01", "Transações de Vendas (recebidas das facilitadoras).", "—", "Não", "Endpoint de ingestão possível via E05; armazenamento/processamento das transações NÃO dimensionado na Fase 1.", "Mesma frente nova de N06/F06.01 — 30M tx/dia, exige camada de dados dedicada (Data Cloud/streaming).")

# ===================================================================
# ABA 3 — REQUISITOS NÃO FUNCIONAIS (6)
# ===================================================================
ws3 = wb.create_sheet("Requisitos Não Funcionais")
style_header(ws3)
r = 2
r = add_group(ws3, r, "6 — Requisitos Não Funcionais")
r = add_row(ws3, r, "RNF-01", "LGPD: Tratamento de dados pessoais (CPF, nome, matrícula) conforme Lei 13.709/2018. Base legal: execução de políticas públicas (Art. 7º, III) e cumprimento de obrigação legal/regulatória (Art. 7º, II). Finalidade documentada; dados coletados limitados ao necessário para operação do programa.", "—", "Integral", "E08 — residência híbrida (ADR 0001): CPF/dados sensíveis não persistem na nuvem Salesforce, resolvidos em runtime; mascaramento de CPF nos logs; diagrama de fluxo sob LGPD.", "")
r = add_row(ws3, r, "RNF-02", "Disponibilidade Módulo de Gestão: 99,5% de segunda a sexta das 7h às 22h.", "—", "Parcial", "Plataforma Salesforce/Hyperforce atende alta disponibilidade nativamente.", "SLA formal (99,5%) não estava acordado no escopo. A validar contra o SLA contratual Salesforce e definir responsabilidade/medição.")
r = add_row(ws3, r, "RNF-03", "Disponibilidade do Marketplace e API: disponíveis 24/7 com SLA reduzido (99%).", "—", "Parcial", "Idem RNF-02 — disponibilidade nativa da plataforma + camada MuleSoft.", "SLA 99% 24/7 formal a validar (inclui MuleSoft on-premise na infra Dataprev — responsabilidade compartilhada).")
r = add_row(ws3, r, "RNF-04", "Tempo de resposta: ≤ 3s para 95% das requisições.", "—", "Parcial", "Arquitetura API-led (E05) e objetos nativos favorecem o alvo.", "NFR formal de performance (≤3s p95) novo — não havia documento de NFR acordado. Impacta desenho de integrações (E05) e portal (E01); a validar/medir.")
r = add_row(ws3, r, "RNF-05", "Capacidade: Suportar até 1 milhão de empresas beneficiárias, até 1 mil facilitadoras, até 1 milhão de estabelecimentos credenciados e até 30 milhões de trabalhadores beneficiados. O módulo de monitoramento deve suportar ingestão de até 30 milhões de transações/dia.", "—", "Parcial", "E08 dimensiona a volumetria de cadastro (~800k estab., ~450k benef.); instância dedicada (ADR 0002).", "Os 30 MILHÕES DE TRANSAÇÕES/DIA do monitoramento NÃO estão dimensionados (frente nova N06). Capacidade de cadastro considerada; capacidade transacional do monitoramento não.")
r = add_row(ws3, r, "RNF-06", "Segurança: Autenticação via gov.br (nível Prata/Ouro) para Beneficiárias e Estabelecimentos; Certificado digital (CNPJ) para facilitadoras e adquirentes via API; GERID MTE para usuários do Ministério; Comunicação entre sistemas (Marketplace e NOVO PAT) protegida por canal seguro.", "—", "Parcial", "E01 — login gov.br (OpenID Connect); E05/E08 — canal seguro entre sistemas.", "NOVAS peças: autenticação por certificado digital/mTLS para facilitadoras e ADQUIRENTES, e GERID MTE como IdP dos usuários do Ministério — a dimensionar em E05/E01.")
r = add_row(ws3, r, "RNF-07", "Auditoria: Todas as ações registradas com timestamp e identificação do usuário, com IP e porta da máquina que originou a operação.", "—", "Integral", "E08 — trilha de auditoria imutável de acesso a dado sensível (observabilidade, Shield/Event Monitoring).", "Registro de IP e PORTA de origem em todas as ações a confirmar no detalhe (granularidade além do padrão de auditoria nativo).")
r = add_row(ws3, r, "RNF-08", "Integração: APIs REST documentadas (OpenAPI 3.0) para facilitadoras e adquirentes.", "—", "Integral", "E05 — camada API-led MuleSoft, contratos documentados (OpenAPI).", "")
r = add_row(ws3, r, "RNF-09", "Acessibilidade: Conformidade com e-MAG 3.1 (Modelo de Acessibilidade em Governo Eletrônico), aderente ao WCAG 2.1 nível AA.", "—", "Parcial", "Experience Cloud (E01) suporta padrões WCAG.", "Conformidade FORMAL e-MAG 3.1 + WCAG 2.1 AA não estava explicitada no escopo — exige verificação/certificação de acessibilidade dos portais. A dimensionar.")
r = add_row(ws3, r, "RNF-10", "Rastreabilidade distribuída: Conformidade com W3C Trace Context (traceparent/tracestate) para correlação de requisições entre serviços. Todos os logs devem incluir o trace-id no formato padronizado.", "—", "Não", "Não previsto no escopo atual.", "NFR novo — propagação de W3C Trace Context entre Salesforce ↔ MuleSoft ↔ sistemas externos e trace-id em todos os logs. A avaliar viabilidade/esforço (E05).")
r = add_row(ws3, r, "RNF-11", "Resiliência: Comunicação com sistemas externos (NOVO PAT, Instituição Financeira) deve prever mecanismos de retry com backoff exponencial e fila de reprocessamento em caso de indisponibilidade.", "—", "Integral", "E05 — mock-first, batch incremental agendado, retry + fila de reprocessamento (Risco 6 do documento: operação não trava por indisponibilidade do Novo PAT).", "Resíduo: menção a 'Instituição Financeira' — as referências à Instituição Bancária foram retiradas na v0.7 (histórico); confirmar se permanece como alvo de integração.")
r = add_row(ws3, r, "RNF-12", "Conformidade com padrões HTTP: Todas as APIs devem seguir rigorosamente a semântica dos códigos de status HTTP conforme RFC 9110 (204 para consulta sem resultado, 404 reservado a recurso inexistente).", "—", "Parcial", "Boa prática já adotada no desenho de APIs MuleSoft (E05).", "Conformidade FORMAL RFC 9110 como requisito contratual é novo; alinhar a padronização de contratos e testar semântica de status.")
r = add_row(ws3, r, "RNF-13", "Volumetria: Trabalhadores beneficiados ~24 milhões; Empresas beneficiárias (por CNPJ) 450 mil; Estabelecimentos ~800 mil (previsão +300 mil até 2030); Facilitadoras ~600 (15 relevantes, 5 dominantes); Transação: volume esperado de 50 milhões/dia.", "—", "Parcial", "Volumetria de cadastro alinhada ao dimensionamento de E08 (instância dedicada).", "O volume TRANSACIONAL (50 mi/dia; monitoramento até 30 mi/dia) refere-se à frente de monitoramento não dimensionada (N06). Informação de contexto para o sizing da nova frente.")

# ===================================================================
# ABA 4 — NÃO ESCOPO DECLARADO PELO DOCUMENTO (3.3)
# ===================================================================
ws4 = wb.create_sheet("Não Escopo (declarado)")
HEADERS4 = ["#", "Item declarado FORA do escopo (MVP nov/2026)", "Nosso alinhamento", "Observação"]
COLW4 = [6, 66, 20, 66]
for c, h in enumerate(HEADERS4, 1):
    cell = ws4.cell(row=1, column=c, value=h)
    cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = wrap_ctr; cell.border = border
for c, w in enumerate(COLW4, 1):
    ws4.column_dimensions[get_column_letter(c)].width = w
ws4.row_dimensions[1].height = 30
ws4.freeze_panes = "A2"

nao_escopo = [
    ("1", "Wallet de governo — utilização da CTPS Digital como meio de pagamento direto.", "Alinhado", "Também fora do nosso escopo."),
    ("2", "Arranjo de pagamentos — operacionalização da transação financeira entre trabalhador e estabelecimento. A plataforma não participa da transação no momento da compra.", "Alinhado (reforça E03/ADR 0003)", "CONFIRMA que a plataforma NÃO faz split transacional em tempo real — é conciliação B2B (folha→boleto→NF). Sustenta recalibrar E03 (hoje XL)."),
    ("3", "Bloqueio de produtos na transação — verificação em tempo real (via código TIP) se o produto é alimentício. Visão de futuro do MTE, não viável para novembro.", "Alinhado", "Fora do nosso escopo."),
    ("4", "Portabilidade ativa — transferência de créditos entre facilitadoras é um arranjo entre elas.", "Alinhado", "Fora do nosso escopo."),
    ("5", "Denúncias e reclamações — canal para trabalhadores registrarem irregularidades via CTPS Digital (hoje via Fala.BR).", "Alinhado", "Fora do nosso escopo (reforça E06 Agentforce deferido)."),
    ("6", "Consulta de saldo do cartão pelo trabalhador — a CTPS Digital informará apenas créditos programados, não saldo disponível.", "Alinhado", "Coerente com F07 (só 'expectativa de crédito' na Fase 1)."),
    ("7", "Infraestrutura pública para facilitadoras — ambiente white-label para facilitadoras menores sem TI própria.", "Alinhado", "Fora do nosso escopo."),
    ("8", "Desenvolvimento de APIs no sistema NOVO PAT — responsabilidade da DTI/MTE. A DATAPREV consumirá as APIs, não as desenvolverá.", "Alinhado (premissa E05)", "Reforça mock-first: Novo PAT hoje sem API; dependência crítica DTI/MTE (Premissa 1 / Risco 4)."),
    ("9", "Gestão de licitações — o processo licitatório em si não é gerido pela plataforma (contratos firmados podem ser registrados).", "Alinhado", "Registro de contratos firmados = F01.07 (integral)."),
    ("10", "Integração com NFC-e da Receita Federal — cruzamento de nota fiscal eletrônica com transações PAT. Versão futura do módulo de monitoramento.", "Alinhado", "Fora do nosso escopo; ligado à frente futura de monitoramento (N06)."),
]
r = 2
for num, item, align, obs in nao_escopo:
    fill = GREEN if align.startswith("Alinhado") else YELLOW
    ws4.cell(row=r, column=1, value=num).alignment = wrap_ctr
    ws4.cell(row=r, column=2, value=item).alignment = wrap_top
    c3 = ws4.cell(row=r, column=3, value=align); c3.alignment = wrap_ctr; c3.fill = PatternFill("solid", fgColor=fill); c3.font = Font(bold=True, size=10)
    ws4.cell(row=r, column=4, value=obs).alignment = wrap_top
    for c in range(1, 5):
        ws4.cell(row=r, column=c).border = border
        if c != 3:
            ws4.cell(row=r, column=c).font = cell_font if c != 1 else id_font
    r += 1

# ===================================================================
# ABA 5 — LEGENDA / RESUMO
# ===================================================================
ws5 = wb.create_sheet("Legenda & Resumo", 0)  # primeira aba
ws5.column_dimensions["A"].width = 4
ws5.column_dimensions["B"].width = 20
ws5.column_dimensions["C"].width = 90
ws5.sheet_view.showGridLines = False

def put(row, col, val, font=None, fill=None, align=None):
    c = ws5.cell(row=row, column=col, value=val)
    if font: c.font = font
    if fill: c.fill = fill
    if align: c.alignment = align
    return c

put(2, 2, "Matriz Comparativa de Requisitos", Font(bold=True, size=16, color=NAVY))
put(3, 2, "Documento de Visão DM.210824 — Implantar MVP para PAT · v0.7 (10/08/2026) · Ministério do Trabalho e Emprego / DATAPREV", Font(size=10, italic=True, color="555555"))
put(4, 2, "Confronto com o escopo Salesforce PS (épicas E01–E10, ADRs 0001–0006). Requisitos transcritos fielmente do documento.", Font(size=10, color="555555"))
put(5, 2, "Gerado em 14/08/2026 · artefato de análise — NÃO altera o projeto (data/*.json intactos).", Font(size=9, italic=True, color="999999"))

put(7, 2, "Legenda da coluna 'Consideramos?'", Font(bold=True, size=12, color=NAVY))
leg = [
    ("Integral", GREEN, "Requisito coberto por inteiro no escopo atual — a coluna 'Como consideramos' indica a épica e a abordagem."),
    ("Parcial", YELLOW, "Coberto com ressalva: divergência de premissa, item não explícito, ou NFR formal novo a validar/dimensionar."),
    ("Não", RED, "Fora do escopo atual / frente nova / conflito com premissa — exige dimensionamento ou decisão antes de assumir."),
]
rr = 8
for label, color, desc in leg:
    c = put(rr, 2, label, Font(bold=True, size=10), PatternFill("solid", fgColor=color), wrap_ctr)
    c.border = border
    put(rr, 3, desc, cell_font, align=wrap_top)
    ws5.row_dimensions[rr].height = 28
    rr += 1

put(12, 2, "Contagem por status", Font(bold=True, size=12, color=NAVY))
# contagem funcional + AD + RNF
counts = {"Integral": 0, "Parcial": 0, "Não": 0}
# preenchida abaixo dinamicamente ao varrer as abas
for sheet in [ws, ws2, ws3]:
    for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4):
        v = row[0].value
        if v in counts:
            counts[v] += 1
rr = 13
total = sum(counts.values())
for k in ["Integral", "Parcial", "Não"]:
    c = put(rr, 2, k, Font(bold=True, size=10), PatternFill("solid", fgColor=STATUS_FILL[k]), wrap_ctr); c.border = border
    put(rr, 3, f"{counts[k]} requisito(s)  ({counts[k]*100//total}%)", cell_font, align=Alignment(vertical="center"))
    rr += 1
put(rr, 2, "TOTAL", Font(bold=True, size=10), PatternFill("solid", fgColor=GREY), wrap_ctr)
put(rr, 3, f"{total} requisitos avaliados (Funcionais + Arquivos de Dados + Não Funcionais)", Font(bold=True, size=10), align=Alignment(vertical="center"))

put(rr+2, 2, "Principais pontos de atenção", Font(bold=True, size=12, color=NAVY))
atencao = [
    "N06 Monitoramento entrou no MVP (até 30M transações/dia) — a proposta o tratava como fora desta rodada. Frente NOVA a dimensionar (staging/streaming, possível Data Cloud).",
    "F02.02 — o documento pede registrar folha 'com trabalhadores e valores'; nosso escopo persiste só o cabeçalho na Fase 1 (linhas = roadmap). Alinhar.",
    "NFRs formais novos: SLA (99%/99,5%), ≤3s p95, W3C Trace Context (RNF-10), RFC 9110 (RNF-12), e-MAG 3.1/WCAG AA (RNF-09) — não havia NFR acordado.",
    "Segurança (RNF-06): certificado digital/mTLS para facilitadoras e adquirentes + GERID MTE como IdP dos usuários do Ministério — peças novas em E05/E01.",
    "Não-escopo item 2 confirma conciliação B2B (sem split transacional em tempo real) — sustenta recalibrar E03 (hoje XL).",
    "Módulo de Gestão MTE (N05) não é épica separada — coberto por Tableau (E10) + UI nativa; delimitar o módulo administrativo.",
]
rr += 3
for a in atencao:
    put(rr, 2, "•", Font(bold=True, size=11, color=BLUE), align=wrap_ctr)
    put(rr, 3, a, cell_font, align=wrap_top)
    ws5.row_dimensions[rr].height = 28
    rr += 1

import os
out = "/Users/nfilho/claude/Scopezilla/DATAPREV-PAT/outputs/artifacts/Matriz_Requisitos_DocVisao_v0.7_vs_Escopo.xlsx"
wb.save(out)
print("OK ->", out)
print("Contagem:", counts, "total", total)
